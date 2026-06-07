from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from datetime import datetime
import csv, io, hashlib
from database import get_db
from auth import get_usuario_atual
import models
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook=None

router=APIRouter(prefix="/importacao", tags=["Importação de Extratos"])

MAP_CATS=[('ifood|mercado|padaria|restaurante|delivery|entrega','Alimentação'),('uber|99|posto|combustivel|metro|ônibus|onibus','Transporte'),('farmacia|drogaria|hospital|clinica','Saúde'),('netflix|spotify|cinema|lazer','Lazer'),('salario|pix recebido|transferencia recebida','Salário')]

def categoria(desc):
    d=(desc or '').lower()
    for termos,cat in MAP_CATS:
        if any(t in d for t in termos.split('|')): return cat
    return 'Outros'

def parse_date(v):
    if isinstance(v, datetime): return v
    s=str(v).strip()
    for fmt in ('%d/%m/%Y','%Y-%m-%d','%d-%m-%Y','%d/%m/%y'):
        try: return datetime.strptime(s[:10], fmt)
        except Exception: pass
    return datetime.utcnow()

def norm_val(v):
    if isinstance(v,(int,float)): return float(v)
    s=str(v).replace('R$','').replace('.','').replace(',','.').strip()
    try: return float(s)
    except Exception: return 0.0

def row_hash(usuario_id, data, desc, valor):
    return hashlib.sha1(f"{usuario_id}|{data.date()}|{desc}|{valor:.2f}".encode()).hexdigest()

def read_csv(content):
    text=content.decode('utf-8-sig', errors='ignore')
    sample=text[:2048]
    dialect=csv.Sniffer().sniff(sample, delimiters=';,\t,')
    return list(csv.DictReader(io.StringIO(text), dialect=dialect))

def read_xlsx(content):
    if not load_workbook: raise HTTPException(400,'Instale openpyxl para importar XLSX')
    wb=load_workbook(io.BytesIO(content), data_only=True); ws=wb.active
    rows=list(ws.iter_rows(values_only=True)); headers=[str(x).strip() for x in rows[0]]
    return [dict(zip(headers,r)) for r in rows[1:] if any(r)]

def normalize(rows, usuario_id, db):
    out=[]
    for r in rows:
        keys={k.lower().strip():k for k in r.keys()}
        data=r.get(keys.get('data') or keys.get('date') or keys.get('dt') or next(iter(r.keys())))
        desc=r.get(keys.get('descrição') or keys.get('descricao') or keys.get('histórico') or keys.get('historico') or keys.get('description') or next(iter(r.keys())))
        val=r.get(keys.get('valor') or keys.get('amount') or keys.get('vlr') or '')
        dt=parse_date(data); valor=norm_val(val)
        if valor==0: continue
        tipo='Receita' if valor>0 else 'Despesa'; valor_abs=abs(valor)
        h=row_hash(usuario_id, dt, desc, valor)
        dup=db.query(models.Movimentacao).filter(models.Movimentacao.usuario_id==usuario_id, models.Movimentacao.hash_importacao==h).first() is not None
        out.append({'data_movimentacao':dt.isoformat(),'descricao':str(desc or ''),'valor':valor_abs,'tipo':tipo,'categoria_sugerida':categoria(str(desc or '')),'hash_importacao':h,'duplicado':dup})
    return out

@router.post('/preview')
def preview(file:UploadFile=File(...), db:Session=Depends(get_db), usuario_atual:models.Usuario=Depends(get_usuario_atual)):
    content=file.file.read(); name=file.filename.lower()
    rows=read_csv(content) if name.endswith('.csv') else read_xlsx(content) if name.endswith('.xlsx') else None
    if rows is None: raise HTTPException(400,'Envie um arquivo CSV ou XLSX')
    return {'arquivo':file.filename,'total_linhas':len(rows),'movimentacoes':normalize(rows, usuario_atual.id, db)}

@router.post('/confirmar')
def confirmar(itens:list[dict], db:Session=Depends(get_db), usuario_atual:models.Usuario=Depends(get_usuario_atual)):
    salvos=0; ignorados=0
    for item in itens:
        if item.get('duplicado') or db.query(models.Movimentacao).filter(models.Movimentacao.usuario_id==usuario_atual.id, models.Movimentacao.hash_importacao==item.get('hash_importacao')).first():
            ignorados+=1; continue
        db.add(models.Movimentacao(usuario_id=usuario_atual.id,tipo=models.TipoMovimentacao.receita if item['tipo']=='Receita' else models.TipoMovimentacao.despesa,valor=float(item['valor']),categoria=item.get('categoria') or item.get('categoria_sugerida') or 'Outros',descricao=item.get('descricao'),data_movimentacao=datetime.fromisoformat(item['data_movimentacao']),hash_importacao=item.get('hash_importacao'))); salvos+=1
    db.commit(); return {'salvos':salvos,'ignorados':ignorados}
